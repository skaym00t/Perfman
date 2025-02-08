import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from datetime import datetime, timedelta
import json
import os
import re

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Менеджер производительности by v.lazarenko")
        
        # Инициализация переменных
        self.total_value = 0
        self.button_history = []
        self.settings = Settings()
        self.timer1 = Timer()
        self.timer2 = Timer()
        self.button_manager = ButtonManager(self)
        self.mode_manager = ModeManager(self)
        self.file_handler = FileHandler()

        # Загрузка настроек
        self.settings.load_settings()
        
        # Создание интерфейса
        self.create_tabs()

        # Создание меню
        self.create_menu()

    def create_tabs(self):
        # Создание вкладок
        notebook = ttk.Notebook(self.root)
        notebook.pack(expand=True, fill='both')

        # Вкладка "Ввод данных"
        input_data_frame = ttk.Frame(notebook)
        notebook.add(input_data_frame, text="Ввод данных")
        self.create_input_data_tab(input_data_frame)

        # Вкладка "Текущий день"
        current_day_frame = ttk.Frame(notebook)
        notebook.add(current_day_frame, text="Текущий день")
        self.create_current_day_tab(current_day_frame)

    def create_input_data_tab(self, frame):
        # Добавление полей для ввода данных
        ttk.Label(frame, text="Логин").grid(row=1, column=0, padx=5, pady=5)
        self.login_entry = ttk.Entry(frame)
        self.login_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame, text="Всего рабочих дней").grid(row=2, column=0, padx=5, pady=5)
        self.total_days_entry = ttk.Entry(frame)
        self.total_days_entry.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(frame, text="Осталось рабочих дней").grid(row=3, column=0, padx=5, pady=5)
        self.remaining_days_entry = ttk.Entry(frame)
        self.remaining_days_entry.grid(row=3, column=1, padx=5, pady=5)

        ttk.Button(frame, text="Загрузить данные", command=self.load_data).grid(row=4, column=0, columnspan=2, padx=5, pady=5)

    def create_current_day_tab(self, frame):
        # Добавление полей для отображения значений
        ttk.Label(frame, text="Кэф на конец месяца/Бонус(в баллах):").grid(row=0, column=0, padx=5, pady=5)
        self.index_value_month_label = ttk.Label(frame, text="0")
        self.index_value_month_label.grid(row=0, column=1, padx=5, pady=5)
        self.bonus_label = ttk.Label(frame, text="0")
        self.bonus_label.grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(frame, text="До следующего бонуса/Следующий бонус(в баллах):").grid(row=2, column=0, padx=5, pady=5)
        self.bonus_goal_label = ttk.Label(frame, text="0")
        self.bonus_goal_label.grid(row=2, column=1, padx=5, pady=5)
        self.next_bonus_label = ttk.Label(frame, text="0")
        self.next_bonus_label.grid(row=2, column=2, padx=5, pady=5)
        
        ttk.Label(frame, text="До нормы(месяц):").grid(row=3, column=0, padx=5, pady=5)
        self.monthly_goal_label = ttk.Label(frame, text="0")
        self.monthly_goal_label.grid(row=3, column=1, padx=5, pady=5)

        ttk.Label(frame, text="До нормы(день):").grid(row=4, column=0, padx=5, pady=5)
        self.daily_goal_label = ttk.Label(frame, text="0")
        self.daily_goal_label.grid(row=4, column=1, padx=5, pady=5)

        current_date = datetime.now().strftime("%d.%m.%Y")  # Получаем текущую дату
        ttk.Label(frame, text=f"Сделано {current_date}:").grid(row=5, column=0, padx=5, pady=5)
        self.done_today_label = ttk.Label(frame, text="0")
        self.done_today_label.grid(row=5, column=1, padx=5, pady=5)
        
        self.index_value_month = 0
        self.bonus = 0
        self.bonus_goal = 0
        self.next_bonus = 0
        self.monthly_goal = 0
        self.daily_goal = 0
        self.done_today = 0
        self.sum_points = 0
        self.norm_value_month = 0
        self.value_now = 0
        self.login = ''

        # Выделенное поле для активностей
        activities_frame = ttk.LabelFrame(frame, text="Активности")
        activities_frame.grid(row=6, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Добавление интерфейса кнопок
        self.button_manager.create_buttons(activities_frame)

        # Кнопка "Отменить последнее действие"
        ttk.Button(frame, text="Отменить последнее действие", command=self.button_manager.undo_last_action).grid(row=7, column=0, columnspan=2, padx=5, pady=5)

        # Выделенное поле для режимов
        modes_frame = ttk.LabelFrame(frame, text="Режимы")
        modes_frame.grid(row=8, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Добавление интерфейса кнопок с таймерами
        self.mode_manager.create_modes(modes_frame)

    # Загрузка данных из файла xlsx
    def load_data(self):
        login = self.login_entry.get()
        total_days = int(self.total_days_entry.get()) if self.total_days_entry.get().isdigit() else None
        remaining_days = int(self.remaining_days_entry.get()) if self.remaining_days_entry.get().isdigit() else None
        if not login or total_days is None or remaining_days is None:
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля корректными данными.")
            return
        try:
            data = self.file_handler.read_from_xlsx(login, total_days, remaining_days)
            if data:
                # Обновляем значения полей до нормы
                self.login = data['login']
                self.daily_goal = data['daily_goal'] - self.done_today
                self.monthly_goal = data['monthly_goal'] - self.done_today
                self.bonus = data['bonus']
                self.bonus_goal = data['bonus_goal'] - self.done_today
                self.sum_points = data['sum_points'] + self.done_today
                self.norm_value_month = data['norm_value_month']
                self.index_value_month = float(self.sum_points / self.norm_value_month) if self.norm_value_month > 0 else 0
                self.next_bonus = data['next_bonus']
                # Обновляем метки на интерфейсе
                self.update_goals_labels(self.daily_goal, self.monthly_goal, data['bonus'], self.bonus_goal, self.index_value_month, data['next_bonus'])
            else:
                messagebox.showerror("Ошибка", "Данные не найдены для указанного логина.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при загрузке данных: {str(e)}")
            
    def update_goals_labels(self, daily_goal, monthly_goal, bonus, bonus_goal, index_value_month, next_bonus):
        self.daily_goal_label.config(text=f"{max(0, int(daily_goal))}")
        self.monthly_goal_label.config(text=f"{max(0, int(monthly_goal))}")
        self.index_value_month_label.config(text=(f"{(index_value_month)}")[:5])
        self.bonus_label.config(text=f"{max(0, int(bonus))}")
        self.bonus_goal_label.config(text=f"{max(0, int(bonus_goal))}")
        self.next_bonus_label.config(text=f"{max(0, int(next_bonus))}")

    # Создание меню
    def create_menu(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        settings_menu = tk.Menu(menu_bar, tearoff=0)
        settings_menu.add_command(label="Добавить новую кнопку", command=self.open_add_button_window)
        settings_menu.add_command(label="Удалить существующую кнопку", command=self.open_remove_button_window)
        settings_menu.add_separator()
        settings_menu.add_command(label="Добавить новый режим", command=self.open_add_mode_window)
        settings_menu.add_command(label="Удалить существующий режим", command=self.open_remove_mode_window)
        settings_menu.add_separator()
        settings_menu.add_command(label="Сброс настроек", command=self.reset_settings)
        settings_menu.add_command(label="Сохранить изменения", command=self.save_settings_changes)
        menu_bar.add_cascade(label="Настройки", menu=settings_menu)

        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Сохранить данные в XLSX", command=self.save_data_to_xlsx)
        menu_bar.add_cascade(label="Файл", menu=file_menu)

    # Открытие окна для добавления новой кнопки
    def open_add_button_window(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Добавить новую кнопку")

        ttk.Label(add_window, text="Название кнопки").grid(row=0, column=0, padx=5, pady=5)
        name_entry = ttk.Entry(add_window)
        name_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(add_window, text="Значение").grid(row=1, column=0, padx=5, pady=5)
        value_entry = ttk.Entry(add_window)
        value_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(add_window, text="Добавить", command=lambda: self.add_button_and_update(name_entry.get(), float(value_entry.get()), add_window)).grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    # Добавление новой кнопки и обновление интерфейса
    def add_button_and_update(self, name, value, window):
        if self.mode_manager.is_any_mode_active():
            messagebox.showwarning("Предупреждение", "Нельзя добавить кнопку, пока включен режим.")
            return
        if len(self.settings.buttons) >= 10:
            messagebox.showwarning("Предупреждение", "Превышено максимальное количество кнопок (10).")
            return
        self.settings.add_button(name, value)
        self.button_manager.reload_buttons()
        window.destroy()
    
    #  Добавление функции удаления кнопки
    def open_remove_button_window(self):
        remove_window = tk.Toplevel(self.root)
        remove_window.title("Удалить кнопку")
        
        ttk.Label(remove_window, text="Выберите кнопку для удаления").grid(row=0, column=0, padx=5, pady=5)
        
        button_var = tk.StringVar(remove_window)
        button_options = list(self.settings.buttons.keys())
        button_dropdown = ttk.OptionMenu(remove_window, button_var, None, *button_options)
        button_dropdown.grid(row=1, column=0, padx=5, pady=5)
        
        ttk.Button(remove_window, text="Удалить", command=lambda: self.remove_button_and_update(button_var.get(), remove_window)).grid(row=2, column=0, padx=5, pady=5)

    def remove_button_and_update(self, button_name, window):
        if button_name:
            self.settings.remove_button(button_name)
            self.button_manager.reload_buttons()
            messagebox.showinfo("Успех", f"Кнопка '{button_name}' удалена.")
            window.destroy()
        else:
            messagebox.showwarning("Предупреждение", "Выберите кнопку для удаления.")

    # Сброс настроек
    def reset_settings(self):
        self.settings.reset_settings()
        self.button_manager.reload_buttons()
        self.mode_manager.reload_modes()
        messagebox.showinfo("Сброс настроек", "Настройки успешно сброшены!")

    # Сохранение изменений настроек
    def save_settings_changes(self):
        self.settings.save_settings()
        self.mode_manager.save_modes()
        messagebox.showinfo("Сохранение изменений", "Изменения успешно сохранены.")
    
    def collect_current_day_data(self):
        # Собираем данные с меток на вкладке "Текущий день"
        login = self.login
        monthly_goal = self.monthly_goal_label.cget("text")
        daily_goal = self.daily_goal_label.cget("text")
        done_today = self.done_today_label.cget("text")
        index_value_month = self.index_value_month_label.cget("text")
        bonus = self.bonus_label.cget("text")
        bonus_goal = self.bonus_goal_label.cget("text")
        next_bonus = self.next_bonus_label.cget("text")
        sum_points = self.sum_points

        # Собираем данные со всех кнопок активностей
        activities_data = []
        for name, button_data in self.button_manager.buttons.items():
            counter = button_data['counter_label'].cget("text")
            cumulative = button_data['cumulative']
            activities_data.append({
                'name': name,
                'counter': counter,
                'cumulative': cumulative
            })
        # Собираем данные со всех режимов
        modes_data = []
        for name, mode_data in self.mode_manager.modes.items():
            timer = mode_data['timer'].visible_timer_label.cget("text")
            modes_data.append({
                'name': name,
                'timer': timer,
            })

        # Возвращаем все собранные данные в виде словаря
        return {
            'login': login,
            'monthly_goal': monthly_goal,
            'daily_goal': daily_goal,
            'done_today': done_today,
            'index_value_month': index_value_month,
            'bonus': bonus,
            'bonus_goal': bonus_goal,
            'next_bonus': next_bonus,
            'sum_points': sum_points,
            'activities': activities_data,
            'modes': modes_data
        }
        
    # Сохранение данных в файл xlsx
    def save_data_to_xlsx(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                initialfile=f"report_{datetime.now().strftime('%d.%m.%Y')}.xlsx")
        if filename:
            # Собираем данные с вкладки "Текущий день"
            data = self.collect_current_day_data()

            # Сохраняем данные в файл XLSX
            wb = openpyxl.Workbook()
            ws = wb.active

            # Записываем заголовки
            cur_day = datetime.now().strftime('%d.%m.%Y')
            headers = [
                "Логин",
                "Месячный план",
                "Дневной план",
                f"Сделано {cur_day}",
                "Кэф на конец месяцa",
                "Бонус за перевыполнение",
                "Баллов до следующего бонуса",
                "Следующий бонус",
                "Всего баллов(по менеджеру)"
                       ]
            ws.append(headers)

            # Записываем значения
            ws.append([
                data['login'],
                data['monthly_goal'],
                data['daily_goal'],
                data['done_today'],
                data['index_value_month'],
                data['bonus'],
                data['bonus_goal'],
                data['next_bonus'],
                data['sum_points']
                ])
            
            #Пустая строка
            ws.append([])

            # Записываем заголовки для активностей
            ws.append(["Название активности", "Количество", "Накопительное значение"])

            # Записываем данные по каждой активности
            for activity in data['activities']:
                ws.append([activity['name'], activity['counter'], activity['cumulative']])
            
            #Пустая строка
            ws.append([])

            # Записываем заголовки для режимов
            ws.append(["Название режима", "Времени в режиме"])

            # Записываем данные по каждому режиму
            for mode in data['modes']:
                ws.append([mode['name'], mode['timer']])

            # Сохраняем файл
            wb.save(filename)
            messagebox.showinfo("Сохранение данных", f"Данные сохранены в {filename}")

    # Открытие окна для добавления нового режима
    def open_add_mode_window(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Добавить новый режим")

        ttk.Label(add_window, text="Название режима").grid(row=0, column=0, padx=5, pady=5)
        name_entry = ttk.Entry(add_window)
        name_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(add_window, text="Баллов в минуту").grid(row=1, column=0, padx=5, pady=5)
        value_entry = ttk.Entry(add_window)
        value_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(add_window, text="Добавить", command=lambda: self.add_mode_and_update(name_entry.get(), value_entry.get(), add_window)).grid(row=2, column=0, columnspan=2, padx=5, pady=5)

    # Добавление нового режима и обновление интерфейса
    def add_mode_and_update(self, name, value_str, window):
        if self.mode_manager.is_any_mode_active():
            messagebox.showwarning("Предупреждение", "Нельзя добавить режим, пока включен режим.")
            return
        value = int(value_str) if value_str.isdigit() else 0
        if len(self.settings.modes) >= 6:
            messagebox.showwarning("Предупреждение", "Превышено максимальное количество режимов (6).")
            return
        self.settings.add_mode(name, value)
        self.mode_manager.reload_modes()
        window.destroy()
    
    def open_remove_mode_window(self):
        remove_window = tk.Toplevel(self.root)
        remove_window.title("Удалить режим")
        
        ttk.Label(remove_window, text="Выберите режим для удаления").grid(row=0, column=0, padx=5, pady=5)
        
        mode_var = tk.StringVar(remove_window)
        mode_options = list(self.settings.modes.keys())
        mode_dropdown = ttk.OptionMenu(remove_window, mode_var, None, *mode_options)
        mode_dropdown.grid(row=1, column=0, padx=5, pady=5)
        
        ttk.Button(remove_window, text="Удалить", command=lambda: self.remove_mode_and_update(mode_var.get(), remove_window)).grid(row=2, column=0, padx=5, pady=5)

    def remove_mode_and_update(self, mode_name, window):
        if mode_name:
            # Проверка наличия активных режимов
            if self.mode_manager.is_any_mode_active():
                messagebox.showwarning("Предупреждение", "Нельзя удалить режим, пока включен режим.")
                return
            
            self.settings.remove_mode(mode_name)
            self.mode_manager.reload_modes()
            
            # Удаляем таймер и кнопку из словарей
            if mode_name in self.mode_manager.timers:
                del self.mode_manager.timers[mode_name]
            if mode_name in self.mode_manager.modes:
                del self.mode_manager.modes[mode_name]
            
            messagebox.showinfo("Успех", f"Режим '{mode_name}' удален.")
            window.destroy()
        else:
            messagebox.showwarning("Предупреждение", "Выберите режим для удаления.")

class Settings:
    SETTINGS_FILE = ".settings.json"

    def __init__(self):
        self.buttons = {}
        self.modes = {}

    # Загрузка настроек из файла
    def load_settings(self):
        if os.path.exists(self.SETTINGS_FILE):
            with open(self.SETTINGS_FILE, 'r') as file:
                data = json.load(file)
                self.buttons = data.get('buttons', {})
                self.modes = data.get('modes', {})

    # Сохранение настроек в файл
    def save_settings(self):
        with open(self.SETTINGS_FILE, 'w') as file:
            json.dump({'buttons': self.buttons, 'modes': self.modes}, file)

    # Сброс настроек
    def reset_settings(self):
        self.buttons = {}
        self.modes = {}
        self.save_settings()

    # Добавление новой кнопки
    def add_button(self, name, value):
        self.buttons[name] = value
        self.save_settings()
        
    # Удаление кнопки
    def remove_button(self, name):
        if name in self.buttons:
            del self.buttons[name]
            self.save_settings()

    # Удаление режима
    def remove_mode(self, name):
        if name in self.modes:
            del self.modes[name]
            self.save_settings()

    # Добавление нового режима
    def add_mode(self, name, value):
        self.modes[name] = value
        self.save_settings()


class Timer:
    def __init__(self):
        self.running = False
        self.start_time = None
        self.elapsed_time = timedelta(seconds=0)
        self.visible_timer_label = None

    # Запуск таймера
    def start_timer(self):
        if not self.running:
            self.start_time = datetime.now()
            self.running = True
            self.update_visible_timer()

    # Остановка таймера
    def stop_timer(self):
        if self.running:
            self.elapsed_time += datetime.now() - self.start_time
            self.running = False

    # Переключение состояния таймера
    def toggle_timer(self):
        if self.running:
            self.stop_timer()
        else:
            self.start_timer()

    # Сброс таймера
    def reset_timer(self):
        self.start_time = datetime.now()
        self.elapsed_time = timedelta(seconds=0)
        if self.visible_timer_label:
            self.visible_timer_label.config(text="0:00:00")

    # Обновление видимого таймера каждую секунду
    def update_visible_timer(self):
        if self.running:
            elapsed = self.elapsed_time + (datetime.now() - self.start_time)
            time_str = str(elapsed).split('.')[0]
            if self.visible_timer_label:
                self.visible_timer_label.config(text=time_str)
            self.visible_timer_label.after(1000, self.update_visible_timer)

    # Создание таймера в интерфейсе
    def create_timer(self, frame, row, col):
        self.visible_timer_label = ttk.Label(frame, text="0:00:00")
        self.visible_timer_label.grid(row=row, column=col, padx=5, pady=5)

        ttk.Button(frame, text="Сброс", command=self.reset_timer).grid(row=row+2, column=col, padx=5, pady=5)


class ButtonManager:
    def __init__(self, app):
        self.app = app
        self.buttons = {}

    # Создание кнопок в поле "Активности"
    def create_buttons(self, frame):
        self.frame = frame  # Сохраняем ссылку на фрейм для перезагрузки кнопок
        self.reload_buttons()

    # Перезагрузка кнопок после изменений настроек
    def reload_buttons(self):
        # Очистка существующих кнопок
        for widget in self.frame.winfo_children():
            if isinstance(widget, (ttk.Button, ttk.Label)):
                widget.destroy()

        # Пересоздание кнопок
        row = 0
        col = 0
        for name, value in self.app.settings.buttons.items():
            self.create_button_row(row, col, name, value)
            col += 2
            if col >= 4:  # Размещаем максимум 2 кнопки в строке
                col = 0
                row += 1

    # Создание строки с кнопкой и метками
    def create_button_row(self, row, col, name, value):
        button = ttk.Button(self.frame, text=name, command=lambda n=name, v=value: self.on_button_click(n, v))
        button.grid(row=row, column=col, padx=5, pady=5)
        counter_label = ttk.Label(self.frame, text="0")
        counter_label.grid(row=row, column=col+1, padx=5, pady=5)
        self.buttons[name] = {
            'value': value,
            'button': button,
            'counter_label': counter_label,
            'counter': 0,
            'cumulative': 0
        }

    # Обработка нажатия на кнопку
    def on_button_click(self, name, value):

        button_data = self.buttons[name]
        button_data['counter'] += 1
        button_data['cumulative'] += value
        button_data['counter_label'].config(text=str(button_data['counter']))

        self.app.done_today += value
        self.app.daily_goal -= value
        self.app.monthly_goal -= value
        self.app.sum_points += value
        self.app.index_value_month = self.app.sum_points / self.app.norm_value_month if self.app.norm_value_month > 0 else 0
        if self.app.index_value_month >= 1.0 and self.app.index_value_month <= 1.11:
            self.app.bonus_goal = (self.app.norm_value_month * 1.11) - self.app.sum_points
            self.app.bonus = 0
            self.app.next_bonus = 1000
        elif self.app.index_value_month > 1.1 and self.app.index_value_month <= 1.21:
            self.app.bonus_goal = (self.app.norm_value_month * 1.21) - self.app.sum_points
            self.app.bonus = 1000
            self.app.next_bonus = 1500
        elif self.app.index_value_month > 1.21 and self.app.index_value_month <= 1.31:
            self.app.bonus_goal = (self.app.norm_value_month * 1.31) - self.app.sum_points
            self.app.bonus = 1500
            self.app.next_bonus = 2000
        elif self.app.index_value_month > 1.31 and self.app.index_value_month <= 1.41:
            self.app.bonus_goal = (self.app.norm_value_month * 1.41) - self.app.sum_points
            self.app.bonus = 2000
            self.app.next_bonus = 2500
        elif self.app.index_value_month > 1.41 and self.app.index_value_month <= 1.51:
            self.app.bonus_goal = (self.app.norm_value_month * 1.51) - self.app.sum_points
            self.app.bonus = 2500
            self.app.next_bonus = 3000
        elif self.app.index_value_month > 1.51 and self.app.index_value_month <= 1.630:
            self.app.bonus_goal = (self.app.norm_value_month * 1.630) - self.app.sum_points
            self.app.bonus = 3000
            self.app.next_bonus = 3750
        elif self.app.index_value_month > 1.625:
            self.app.bonus_goal = 0
            self.app.bonus = 3750
            self.app.next_bonus = 'Ты молодец!'
        else:
            self.app.bonus_goal = 0
            self.app.bonus = 0
            self.app.next_bonus = 0      

        self.update_goals()

        self.app.button_history.append((name, value))

    # Обновление полей "До нормы"
    def update_goals(self):
        self.app.daily_goal_label.config(text=f"{max(0, int(self.app.daily_goal))}")
        self.app.monthly_goal_label.config(text=f"{max(0, int(self.app.monthly_goal))}")
        self.app.done_today_label.config(text=f"{int(self.app.done_today)}")
        self.app.index_value_month_label.config(text=(f"{(self.app.index_value_month)}")[:5])
        self.app.bonus_label.config(text=f"{max(0, int(self.app.bonus))}")
        self.app.bonus_goal_label.config(text=f"{max(0, int(self.app.bonus_goal))}")
        self.app.next_bonus_label.config(text=f"{self.app.next_bonus}")
        
    # Отмена последнего действия
    def undo_last_action(self):
        if self.app.button_history:
            last_action = self.app.button_history.pop()
            name, value = last_action

            button_data = self.buttons[name]
            if button_data['counter'] > 0:
                button_data['counter'] -= 1
                button_data['cumulative'] -= value
                button_data['counter_label'].config(text=str(button_data['counter']))

            self.app.done_today -= value
            self.app.daily_goal += value
            self.app.monthly_goal += value
            self.app.sum_points -= value
            self.app.index_value_month = self.app.sum_points / self.app.norm_value_month if self.app.norm_value_month > 0 else 0
            if self.app.index_value_month >= 1.0 and self.app.index_value_month <= 1.11:
                self.app.bonus_goal = (self.app.norm_value_month * 1.11) - self.app.sum_points
                self.app.bonus = 0
                self.app.next_bonus = 1000
            elif self.app.index_value_month > 1.1 and self.app.index_value_month <= 1.21:
                self.app.bonus_goal = (self.app.norm_value_month * 1.21) - self.app.sum_points
                self.app.bonus = 1000
                self.app.next_bonus = 1500
            elif self.app.index_value_month > 1.21 and self.app.index_value_month <= 1.31:
                self.app.bonus_goal = (self.app.norm_value_month * 1.31) - self.app.sum_points
                self.app.bonus = 1500
                self.app.next_bonus = 2000
            elif self.app.index_value_month > 1.31 and self.app.index_value_month <= 1.41:
                self.app.bonus_goal = (self.app.norm_value_month * 1.41) - self.app.sum_points
                self.app.bonus = 2000
                self.app.next_bonus = 2500
            elif self.app.index_value_month > 1.41 and self.app.index_value_month <= 1.51:
                self.app.bonus_goal = (self.app.norm_value_month * 1.51) - self.app.sum_points
                self.app.bonus = 2500
                self.app.next_bonus = 3000
            elif self.app.index_value_month > 1.51 and self.app.index_value_month <= 1.630:
                self.app.bonus_goal = (self.app.norm_value_month * 1.630) - self.app.sum_points
                self.app.bonus = 3000
                self.app.next_bonus = 3750
            elif self.app.index_value_month > 1.625:
                self.app.bonus_goal = 0
                self.app.bonus = 3750
                self.app.next_bonus = 'Ты молодец!'
            else:
                self.app.bonus_goal = 0
                self.app.bonus = 0
                self.app.next_bonus = 0  

            self.update_goals()


class ModeManager:
    def __init__(self, app):
        self.app = app
        self.modes = {}
        self.timers = {}
        self.active_mode = None

    # Создание режимов в поле "Режимы"
    def create_modes(self, frame):
        self.frame = frame  # Сохраняем ссылку на фрейм для перезагрузки режимов
        self.reload_modes()

    # Перезагрузка режимов после изменений настроек
    def reload_modes(self):
        # Очистка существующих режимов
        for widget in self.frame.winfo_children():
            if isinstance(widget, (ttk.Button, ttk.Label, ttk.Frame)):
                widget.destroy()
        
        # Пересоздание режимов
        row = 0
        col = 0
        for name, value in self.app.settings.modes.items():
            self.create_mode_row(row, col, name, value)
            col += 1
            if col >= 3:  # Размещаем максимум 3 кнопки в строке
                col = 0
                row += 1
    
    # Создание строки с кнопкой и таймером
    def create_mode_row(self, row, col, name, value):
        mode_frame = ttk.Frame(self.frame)
        mode_frame.grid(row=row, column=col, padx=5, pady=5)
        timer = Timer()
        self.timers[name] = timer
        # Описание как выглядит кнопка с таймером
        button = ttk.Button(mode_frame, text=name, command=lambda n=name, v=value: self.on_mode_click(n, v, timer))
        button.grid(row=0, column=0, padx=5, pady=5)
        self.modes[name] = {
            'value': value,
            'button': button,
            'timer': timer
        }
        timer.create_timer(mode_frame, row=1, col=0)

    # Обработка нажатия на режим
    def on_mode_click(self, name, value, timer):
        if self.active_mode and self.active_mode == name:
            # Если уже активен этот режим, просто останавливаем его
            timer.toggle_timer()
            if not timer.running:
                self.active_mode = None
                self.enable_all_buttons()
        else:
            # Остановить все остальные режимы и деактивировать их кнопки
            for mode_name, mode_data in self.modes.items():
                if mode_name != name:
                    mode_data['timer'].stop_timer()
                    mode_data['button'].config(state=tk.NORMAL)

            # Активировать выбранный режим
            timer.toggle_timer()
            self.active_mode = name
            self.disable_other_modes(name)

            # Если режим добавляет значение, то запустить таймер для добавления значений
            if value > 0 and timer.running:
                self.start_value_timer(name, value, timer)

    def is_any_mode_active(self):
        for mode_data in self.modes.values():
            if mode_data['timer'].running:
                return True
        return False

    # Запуск таймера для добавления значений
    def start_value_timer(self, name, value, timer):
        if timer.running:
            elapsed_seconds = int((timer.elapsed_time + (datetime.now() - timer.start_time)).total_seconds())

            if elapsed_seconds % 60 == 0 and elapsed_seconds > 0:
                self.app.done_today += value
                self.app.daily_goal -= value
                self.app.monthly_goal -= value
                self.app.sum_points += value
                self.app.index_value_month = self.app.sum_points / self.app.norm_value_month if self.app.norm_value_month > 0 else 0
                if self.app.index_value_month >= 1.0 and self.app.index_value_month <= 1.11:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.11) - self.app.sum_points
                    self.app.bonus = 0
                    self.app.next_bonus = 1000
                elif self.app.index_value_month > 1.1 and self.app.index_value_month <= 1.21:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.21) - self.app.sum_points
                    self.app.bonus = 1000
                    self.app.next_bonus = 1500
                elif self.app.index_value_month > 1.21 and self.app.index_value_month <= 1.31:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.31) - self.app.sum_points
                    self.app.bonus = 1500
                    self.app.next_bonus = 2000
                elif self.app.index_value_month > 1.31 and self.app.index_value_month <= 1.41:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.41) - self.app.sum_points
                    self.app.bonus = 2000
                    self.app.next_bonus = 2500
                elif self.app.index_value_month > 1.41 and self.app.index_value_month <= 1.51:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.51) - self.app.sum_points
                    self.app.bonus = 2500
                    self.app.next_bonus = 3000
                elif self.app.index_value_month > 1.51 and self.app.index_value_month <= 1.630:
                    self.app.bonus_goal = (self.app.norm_value_month * 1.630) - self.app.sum_points
                    self.app.bonus = 3000
                    self.app.next_bonus = 3750
                elif self.app.index_value_month > 1.625:
                    self.app.bonus_goal = 0
                    self.app.bonus = 3750
                    self.app.next_bonus = 'Ты молодец!'
                else:
                    self.app.bonus_goal = 0
                    self.app.bonus = 0
                    self.app.next_bonus = 0  

                self.update_goals()

            self.app.root.after(1000, lambda: self.start_value_timer(name, value, timer))

    # Обновление полей "До нормы"
    def update_goals(self):
        self.app.daily_goal_label.config(text=f"{max(0, int(self.app.daily_goal))}")
        self.app.monthly_goal_label.config(text=f"{max(0, int(self.app.monthly_goal))}")
        self.app.done_today_label.config(text=f"{int(self.app.done_today)}")
        self.app.index_value_month_label.config(text=(f"{(self.app.index_value_month)}")[:5])
        self.app.bonus_label.config(text=f"{max(0, int(self.app.bonus))}")
        self.app.bonus_goal_label.config(text=f"{max(0, int(self.app.bonus_goal))}")
        self.app.next_bonus_label.config(text=f"{self.app.next_bonus}")

    # Деактивация всех других режимов
    def disable_other_modes(self, active_mode):
        for mode_name, mode_data in self.modes.items():
            if mode_name != active_mode:
                mode_data['button'].config(state=tk.DISABLED)
        
        active_mode_value = self.modes.get(active_mode, {}).get('value', 0)

        if active_mode_value > 0 or active_mode_value == None:
            self.disable_activity_buttons()
        else:
            self.enable_activity_buttons()

    # Деактивация кнопок активности
    def disable_activity_buttons(self):
        for widget in self.app.button_manager.frame.winfo_children():
            if isinstance(widget, ttk.Button):
                widget.config(state=tk.DISABLED)

    # Активация кнопок активности
    def enable_activity_buttons(self):
        for widget in self.app.button_manager.frame.winfo_children():
            if isinstance(widget, ttk.Button):
                widget.config(state=tk.NORMAL)

    # Активация всех кнопок
    def enable_all_buttons(self):
        for widget in self.app.button_manager.frame.winfo_children():
            if isinstance(widget, ttk.Button):
                widget.config(state=tk.NORMAL)

        for widget in self.frame.winfo_children():
            if isinstance(widget, ttk.Frame):
                button = widget.winfo_children()[0]
                button.config(state=tk.NORMAL)

    # Сохранение режимов
    def save_modes(self):
        self.app.settings.save_settings()


from datetime import datetime
import re
import openpyxl
import tkinter.messagebox as messagebox

class FileHandler:
    def read_from_xlsx(self, login, total_days, remaining_days):
        try:
            # Поиск файла, содержащего слово "мотивация" в названии
            target_file = None
            for filename in os.listdir('.'):
                if "мотивация" in filename.casefold() and filename.endswith('.xlsx'):  # Регистронезависимая проверка
                    target_file = filename
                    break
            
            # Если файл не найден, выводим сообщение об ошибке
            if not target_file:
                messagebox.showwarning("Предупреждение", "Файл с мотивацией не найден.")
            else:
                    # Загружаем workbook
                    workbook = openpyxl.load_workbook(target_file)
                    sheet = workbook['Данные за период']
                    
                    # Получаем заголовки столбцов
                    headers = [cell.value for cell in sheet[1]]
                    print(f"Заголовки столбцов: {headers}")
                    
                    # Проверяем наличие необходимых столбцов
                    required_columns = ['Логин', 'График', 'Период', 'Сумма баллов']
                    missing_columns = [col for col in required_columns if col not in headers]
                    if missing_columns:
                        messagebox.showerror("Ошибка", f"Отсутствуют необходимые столбцы: {', '.join(missing_columns)}")
                        return None
                    
                    # Находим индексы столбцов
                    login_index = headers.index('Логин')
                    grafik_index = headers.index('График')
                    season_index = headers.index('Период')
                    sum_points_index = headers.index('Сумма баллов')
                    
                    # Преобразуем текущую дату в формат YYYY-MM-DD
                    current_date = datetime.now().strftime("%Y-%m-01")
                    
                    # Переменные для хранения данных из нужной строки
                    target_row = None
                    
                    # Ищем строку с нужным периодом и логином
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if current_date in str(row[season_index]) and row[login_index] and row[login_index].lower() == login.lower():
                            target_row = row
                            grafik = target_row[grafik_index]
                            sum_points = target_row[sum_points_index]
                            break
                    
                    # Если строка с текущим периодом не найдена, ищем строку с указанным логином
                    if not target_row:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if row[login_index] and row[login_index].lower() == login.lower():
                                target_row = row
                                grafik = target_row[grafik_index]
                                sum_points = 0
                                messagebox.showwarning("Предупреждение", "Строка с текущим периодом не найдена или первый день месяца.")
                                break
                    
                    # Если ни одна строка с указанным логином не найдена, прерываем выполнение
                    if not target_row:
                        messagebox.showwarning("Предупреждение", "Логин не найден в файле.")
                        return None
                    
                    # Определяем значение нормы в зависимости от графика
                    if '8' in str(grafik):
                        value_month = 480 * total_days
                        value_now = 480 * (total_days - remaining_days) if total_days != remaining_days else 480
                        norm_value_day = ((480 * total_days) - sum_points) / remaining_days if remaining_days > 0 else 0
                        norm_value_month = (480 * total_days) - sum_points
                        performance = float(sum_points / value_month) if value_month != 0 else 0
                    elif '2/2' in str(grafik):
                        value_month = 620 * total_days
                        value_now = 620 * (total_days - remaining_days) if total_days != remaining_days else 620
                        norm_value_day = ((620 * total_days) - sum_points) / remaining_days if remaining_days > 0 else 0
                        norm_value_month = (620 * total_days) - sum_points
                        performance = float(sum_points / value_month) if value_month != 0 else 0
                    else:
                        messagebox.showwarning("Предупреждение", f"Неизвестное значение графика: {grafik}")
                        return None
                    
                    # Определение бонусов на основе производительности
                    if performance >= 1.0 and performance <= 1.11:
                        norm_value_bonus = (norm_value_month * 1.11) - sum_points
                        sum_bonus = 0
                        next_bonus = 1000
                    elif performance > 1.1 and performance <= 1.21:
                        norm_value_bonus = (norm_value_month * 1.21) - sum_points
                        sum_bonus = 1000
                        next_bonus = 1500
                    elif performance > 1.2 and performance <= 1.31:
                        norm_value_bonus = (norm_value_month * 1.31) - sum_points
                        sum_bonus = 1500
                        next_bonus = 2000
                    elif performance > 1.3 and performance <= 1.41:
                        norm_value_bonus = (norm_value_month * 1.41) - sum_points
                        sum_bonus = 2000
                        next_bonus = 2500
                    elif performance > 1.4 and performance <= 1.51:
                        norm_value_bonus = (norm_value_month * 1.51) - sum_points
                        sum_bonus = 2500
                        next_bonus = 3000
                    elif performance > 1.5 and performance <= 1.630:
                        norm_value_bonus = (norm_value_month * 1.630) - sum_points
                        sum_bonus = 3000
                        next_bonus = 3750
                    elif performance > 1.625:
                        norm_value_bonus = 0
                        sum_bonus = 3750
                        next_bonus = 'Ты молодец!'
                    else:
                        norm_value_bonus = 0
                        sum_bonus = 0 
                        next_bonus = 0
                    
                    # Возвращаем результат
                    return {
                        'login': login,
                        'daily_goal': norm_value_day,
                        'monthly_goal': norm_value_month,
                        'bonus_goal': norm_value_bonus,
                        'bonus': sum_bonus,
                        'index_value_month': performance,
                        'norm_value_month': value_month,
                        'sum_points': sum_points,
                        'value_now': value_now,
                        'next_bonus': next_bonus
                    }
            
            # Если файл не найден
            messagebox.showwarning("Предупреждение", "Файл не найден.")
            return None
        
        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке файла: {e}")
        return None


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()