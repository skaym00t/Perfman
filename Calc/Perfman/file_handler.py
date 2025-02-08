import os
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

class FileHandler:
    def read_from_xlsx(self, login, total_days, remaining_days):
        try:
            root = tk.Tk()
            root.withdraw()  # Скрыть главное окно
            target_file = filedialog.askopenfilename(
                title="Выберите файл с мотивацией",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not target_file:
                messagebox.showwarning("Предупреждение", "Файл с мотивацией не выбран.")
                return None

            workbook = openpyxl.load_workbook(target_file)
            sheet = workbook['Данные за период']
            headers = [cell.value for cell in sheet[1]]
            print(f"Заголовки столбцов: {headers}")
            required_columns = ['Логин', 'График', 'Период', 'Сумма баллов']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                messagebox.showerror("Ошибка", f"Отсутствуют необходимые столбцы: {', '.join(missing_columns)}")
                return None
            login_index = headers.index('Логин')
            grafik_index = headers.index('График')
            season_index = headers.index('Период')
            sum_points_index = headers.index('Сумма баллов')
            current_date = datetime.now().strftime("%Y-%m-01")
            target_row = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if current_date in str(row[season_index]) and row[login_index] and row[login_index].lower() == login.lower():
                    target_row = row
                    grafik = target_row[grafik_index]
                    sum_points = target_row[sum_points_index]
                    break
            if not target_row:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[login_index] and row[login_index].lower() == login.lower():
                        target_row = row
                        grafik = target_row[grafik_index]
                        sum_points = 0
                        messagebox.showwarning("Предупреждение", "Строка с текущим периодом не найдена или первый день месяца.")
                        break
            if not target_row:
                messagebox.showwarning("Предупреждение", "Логин не найден в файле.")
                return None
            if '8' in str(grafik):
                value_month = 480 * total_days
                value_now = 480 * (total_days - remaining_days) if total_days != remaining_days else 480
                norm_value_day = ((480 * total_days) - sum_points) / remaining_days if remaining_days > 0 else 0
                norm_value_month = (480 * total_days) - sum_points
                performance = float(sum_points / value_month) if value_month != 0 else 0
            elif '2/2' in str(grafik):
                value_month = 620 * total_days
                value_now = 620 * (total_days - remaining_days) if total_days != remaining_days else 620
                norm_value_day = ((620 * total_days) - sum_points) / remaining_days if remaining_days > 0 else 0
                norm_value_month = (620 * total_days) - sum_points
                performance = float(sum_points / value_month) if value_month != 0 else 0
            else:
                messagebox.showwarning("Предупреждение", f"Неизвестное значение графика: {grafik}")
                return None
            if performance >= 1.0 and performance <= 1.11:
                norm_value_bonus = (norm_value_month * 1.11) - sum_points
                sum_bonus = 0
                next_bonus = 1000
            elif performance > 1.1 and performance <= 1.21:
                norm_value_bonus = (norm_value_month * 1.21) - sum_points
                sum_bonus = 1000
                next_bonus = 1500
            elif performance > 1.2 and performance <= 1.31:
                norm_value_bonus = (norm_value_month * 1.31) - sum_points
                sum_bonus = 1500
                next_bonus = 2000
            elif performance > 1.3 and performance <= 1.41:
                norm_value_bonus = (norm_value_month * 1.41) - sum_points
                sum_bonus = 2000
                next_bonus = 2500
            elif performance > 1.4 and performance <= 1.51:
                norm_value_bonus = (norm_value_month * 1.51) - sum_points
                sum_bonus = 2500
                next_bonus = 3000
            elif performance > 1.5 and performance <= 1.630:
                norm_value_bonus = (norm_value_month * 1.630) - sum_points
                sum_bonus = 3000
                next_bonus = 3750
            elif performance > 1.625:
                norm_value_bonus = 0
                sum_bonus = 3750
                next_bonus = 'Ты молодец!'
            else:
                norm_value_bonus = 0
                sum_bonus = 0 
                next_bonus = 0
            return {
                'login': login,
                'daily_goal': norm_value_day,
                'monthly_goal': norm_value_month,
                'bonus_goal': norm_value_bonus,
                'bonus': sum_bonus,
                'index_value_month': performance,
                'norm_value_month': value_month,
                'sum_points': sum_points,
                'value_now': value_now,
                'next_bonus': next_bonus
            }
        except FileNotFoundError:
            messagebox.showerror("Ошибка", "Файл не найден.")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке файла: {e}")
        return None
